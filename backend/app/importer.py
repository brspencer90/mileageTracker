"""Reusable fill-up ingestion (MT-3).

`import_fillups()` owns the *sink*; source adapters produce normalized dicts.
The later SQL Server migration is just a second adapter feeding the same sink.

MPG formula note: v2 derives MPG as mpf / *current* row's gallons (standard
full-tank method: gallons purchased now = fuel burned since last fill). The
historical CSV divided by the *previous* row's gallons (verified: row 2 MPG
25.32 = 293/11.571, not 293/11.168), so recomputed historical MPG differs by
roughly 0.5-1.5 from the stored column. That is a correction, not a bug. The
stored MPF/MPG columns are ignored on import; they are only cross-checked
(using the legacy prev-gallons formula) to catch parse errors, appending a
warning when |stored - recomputed| > 0.05.
"""

import csv
import re
import sqlite3
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Iterator, NotRequired, TypedDict

from openpyxl import load_workbook

NormalizedRow = TypedDict("NormalizedRow", {
    "vehicle_name": str, "date": str,        # ISO YYYY-MM-DD
    "mileage": int, "gallons": float,
    "cost": float | None,                    # one historical row has no recorded cost
    "station": str | None, "zip": str | None, "missed_last_fill": bool,
    # MT-20/MT-21 extensions; absent (csv adapter) means False / None.
    "mileage_estimated": NotRequired[bool],
    "gauge_notches": NotRequired[float | None],
})

_NA_VALUES = {"", "#N/A"}
_ZIP_RE = re.compile(r"^\d{5}$")

_XLSX_DEFAULT_VEHICLE = "Tiger"
# The xlsx history has a ~10-month unrecorded gap (Nov 2023 -> Sep 2024) whose
# first following row is NOT flagged missed in the source. Any gap this long
# necessarily spans unrecorded fills, so the adapter forces the flag.
_MISSED_GAP_DAYS = 180


@dataclass
class ImportReport:
    inserted: int = 0
    skipped_existing: int = 0
    vehicles_created: int = 0
    skipped_placeholder: int = 0  # populated by callers from the adapter's stats
    estimated: int = 0            # inserted rows carrying mileage_estimated=True
    warnings: list[str] = field(default_factory=list)


def import_fillups(
    conn: sqlite3.Connection,
    rows: Iterable[NormalizedRow],
    dry_run: bool = False,
) -> ImportReport:
    """Insert normalized rows in a single transaction.

    Vehicles are get-or-created by name. Inserts use INSERT OR IGNORE keyed on
    UNIQUE(vehicle_id, mileage), so re-running the same source is naturally
    idempotent. With dry_run=True the transaction is rolled back at the end.
    """
    report = ImportReport()
    vehicle_ids: dict[str, int] = {}
    for row in rows:
        name = row["vehicle_name"]
        if name not in vehicle_ids:
            existing = conn.execute(
                "SELECT id FROM vehicles WHERE name = ?", (name,)
            ).fetchone()
            if existing:
                vehicle_ids[name] = existing[0]
            else:
                cur = conn.execute(
                    "INSERT INTO vehicles (name) VALUES (?)", (name,)
                )
                vehicle_ids[name] = cur.lastrowid
                report.vehicles_created += 1
        cur = conn.execute(
            "INSERT OR IGNORE INTO fillups"
            " (vehicle_id, date, mileage, gallons, cost, station, zip,"
            "  missed_last_fill, mileage_estimated, gauge_notches)"
            " VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                vehicle_ids[name],
                row["date"],
                row["mileage"],
                row["gallons"],
                row["cost"],
                row["station"],
                row["zip"],
                1 if row["missed_last_fill"] else 0,
                1 if row.get("mileage_estimated", False) else 0,
                row.get("gauge_notches"),
            ),
        )
        if cur.rowcount:
            report.inserted += 1
            if row.get("mileage_estimated", False):
                report.estimated += 1
        else:
            report.skipped_existing += 1
    if dry_run:
        conn.rollback()
    else:
        conn.commit()
    return report


def rows_from_legacy_csv(
    path: str | Path, warnings: list[str] | None = None
) -> Iterator[NormalizedRow]:
    """Adapter for the legacy mileageTracker.csv export.

    Handles (all verified against the real file): UTF-8 BOM, case-insensitive
    headers (the real header is 'Zip code', not 'Zip Code'), '#N/A'/empty ->
    None, float-formatted and short zips, MM/DD/YYYY dates, and the stored
    MPF/MPG columns which are ignored (see module docstring).
    """
    if warnings is None:
        warnings = []
    prev: dict[str, tuple[int, float]] = {}  # vehicle -> (mileage, gallons)
    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        reader.fieldnames = [h.strip().lower() for h in reader.fieldnames or []]
        for lineno, raw in enumerate(reader, start=2):
            r = {k: _na(v) for k, v in raw.items()}
            try:
                vehicle = r["car"]
                if not vehicle:
                    raise ValueError("missing Car")
                row: NormalizedRow = {
                    "vehicle_name": vehicle,
                    "date": _parse_date(r["date"]),
                    "mileage": int(float(r["mileage"])),
                    "gallons": float(r["gallons"]),
                    "cost": float(r["cost"]) if r["cost"] is not None else None,
                    "station": r["gas station"],
                    "zip": _norm_zip(r["zip code"]),
                    "missed_last_fill": bool(int(float(r["missed last fill"] or 0))),
                }
            except (KeyError, TypeError, ValueError) as exc:
                warnings.append(f"line {lineno}: skipped unparseable row ({exc})")
                continue
            if row["cost"] is None:
                warnings.append(f"line {lineno}: missing cost imported as NULL")

            # Cross-check stored MPG using the legacy prev-gallons formula to
            # catch parse errors (not the known formula difference).
            stored_mpg = r.get("mpg")
            if (
                stored_mpg is not None
                and not row["missed_last_fill"]
                and vehicle in prev
            ):
                prev_mileage, prev_gallons = prev[vehicle]
                recomputed = (row["mileage"] - prev_mileage) / prev_gallons
                try:
                    if abs(float(stored_mpg) - recomputed) > 0.05:
                        warnings.append(
                            f"line {lineno}: stored MPG {stored_mpg} differs from"
                            f" recomputed {recomputed:.2f} by more than 0.05"
                        )
                except ValueError:
                    warnings.append(f"line {lineno}: unparseable stored MPG {stored_mpg!r}")
            prev[vehicle] = (row["mileage"], row["gallons"])
            yield row


def rows_from_xlsx(
    path: str | Path,
    warnings: list[str] | None = None,
    stats: dict[str, int] | None = None,
) -> list[NormalizedRow]:
    """Adapter for MileageTracker.xlsx (MT-20 / MT-21).

    Policy (validated by the 2026-07-12 investigation, PRODUCT_PLAN Epic 4b):

    - openpyxl ``data_only=True`` (cached formula values, not formulas);
      columns are positional A-I; formula columns J/K are ignored entirely.
    - Blank Car cells default to the last seen name ('Tiger' initially).
    - Blank Missed Last Fill -> 0; the unnamed gauge column H imports raw as
      nullable ``gauge_notches``; int-typed zip cells normalize to 5-char
      strings; blank cost imports as NULL (with a warning), like the csv path.
    - Placeholder rows (a date but mileage+gallons+cost all blank) are
      skipped and counted in ``stats['skipped_placeholder']``.
    - Gap rule: a kept row dated more than ``_MISSED_GAP_DAYS`` after the
      previous kept row for the same vehicle is forced missed_last_fill=True
      (see the constant's comment: the Nov 2023 -> Sep 2024 gap is unflagged
      in the source, and no single tank spans 10 months).
    - Missing-mileage reconstruction (MT-21): a run of real fills with blank
      mileage bracketed by known odometer readings gets gallons-weighted
      interpolated mileage: the bracket's mileage delta is split across the
      intervening intervals proportionally to the gallons purchased at each
      interval's *ending* fill (the closing anchor's gallons is the last
      weight). Estimated rows are flagged ``mileage_estimated=True`` and each
      value is reported as a warning. Unbracketed blank-mileage rows (e.g.
      trailing ones) are skipped with a warning.

    Returns a list (not a generator) so ``warnings``/``stats`` are fully
    populated on return — reconstruction needs the whole sheet anyway.
    """
    if warnings is None:
        warnings = []
    if stats is None:
        stats = {}
    stats.setdefault("skipped_placeholder", 0)

    parsed: list[tuple[int, dict]] = []  # (excel row number, row dict)
    prev_dates: dict[str, date] = {}     # vehicle -> date of last kept row
    vehicle = _XLSX_DEFAULT_VEHICLE
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        for rowno, cells in enumerate(
            ws.iter_rows(min_row=2, max_col=9, values_only=True), start=2
        ):
            cells = tuple(cells) + (None,) * (9 - len(cells))
            car, raw_date, mileage, gallons, cost, station, zip_raw, gauge, missed = cells
            if all(v is None for v in cells):
                continue  # trailing empty row
            if raw_date is not None and mileage is None and gallons is None and cost is None:
                stats["skipped_placeholder"] += 1
                warnings.append(
                    f"row {rowno}: skipped placeholder (date only, no fill data)"
                )
                continue
            if isinstance(car, str) and car.strip():
                vehicle = car.strip()
            try:
                row = {
                    "vehicle_name": vehicle,
                    "date": _parse_xlsx_date(raw_date),
                    "mileage": int(mileage) if mileage is not None else None,
                    "gallons": float(gallons),
                    "cost": float(cost) if cost is not None else None,
                    "station": station.strip()
                    if isinstance(station, str) and station.strip()
                    else None,
                    "zip": _norm_xlsx_zip(zip_raw),
                    "missed_last_fill": bool(int(missed)) if missed is not None else False,
                    "mileage_estimated": False,
                    "gauge_notches": float(gauge) if gauge is not None else None,
                }
            except (TypeError, ValueError) as exc:
                warnings.append(f"row {rowno}: skipped unparseable row ({exc})")
                continue
            if row["cost"] is None:
                warnings.append(f"row {rowno}: missing cost imported as NULL")

            this_date = date.fromisoformat(row["date"])
            prev_date = prev_dates.get(vehicle)
            if prev_date is not None and not row["missed_last_fill"]:
                gap = (this_date - prev_date).days
                if gap > _MISSED_GAP_DAYS:
                    row["missed_last_fill"] = True
                    warnings.append(
                        f"row {rowno}: forced missed_last_fill=1"
                        f" ({gap} days since the previous recorded fill)"
                    )
            prev_dates[vehicle] = this_date
            parsed.append((rowno, row))
    finally:
        wb.close()

    dropped = _reconstruct_missing_mileage(parsed, warnings)
    return [row for i, (_, row) in enumerate(parsed) if i not in dropped]


def _reconstruct_missing_mileage(
    parsed: list[tuple[int, dict]], warnings: list[str]
) -> set[int]:
    """Fill blank-mileage runs in place via gallons-weighted interpolation.

    Returns the set of ``parsed`` indices to drop (unbracketed blank rows).
    """
    by_vehicle: dict[str, list[int]] = {}
    for i, (_, row) in enumerate(parsed):
        by_vehicle.setdefault(row["vehicle_name"], []).append(i)

    dropped: set[int] = set()
    for idxs in by_vehicle.values():
        j = 0
        while j < len(idxs):
            if parsed[idxs[j]][1]["mileage"] is not None:
                j += 1
                continue
            k = j  # run of consecutive blank-mileage rows: idxs[j:k]
            while k < len(idxs) and parsed[idxs[k]][1]["mileage"] is None:
                k += 1
            before = parsed[idxs[j - 1]][1] if j > 0 else None
            after = parsed[idxs[k]][1] if k < len(idxs) else None
            run = [parsed[idxs[m]] for m in range(j, k)]
            if before is None or before["mileage"] is None or after is None:
                for rowno, _ in run:
                    warnings.append(
                        f"row {rowno}: skipped fill with blank mileage"
                        " (no bracketing odometer readings to interpolate)"
                    )
                dropped.update(idxs[m] for m in range(j, k))
                j = k
                continue
            delta = after["mileage"] - before["mileage"]
            # Weight each interval by the gallons purchased at its ENDING
            # fill (fuel bought now ~ miles driven since the previous fill);
            # the closing anchor's gallons is the last weight.
            weights = [row["gallons"] for _, row in run] + [after["gallons"]]
            total = sum(weights)
            cum = 0.0
            for (rowno, row), weight in zip(run, weights):
                cum += weight
                estimate = round(before["mileage"] + delta * cum / total)
                row["mileage"] = estimate
                row["mileage_estimated"] = True
                warnings.append(
                    f"row {rowno}: estimated mileage {estimate}"
                    " via gallons-weighted interpolation"
                )
            j = k
    return dropped


def _parse_xlsx_date(value) -> str:
    """openpyxl gives real datetime/date objects for date-formatted cells."""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raise ValueError(f"missing or non-date Date cell: {value!r}")


def _norm_xlsx_zip(value) -> str | None:
    """xlsx zip cells are int-typed (occasionally float); route through the
    same 5-digit normalization as the csv adapter."""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return _norm_zip(str(value))


def _na(value: str | None) -> str | None:
    if value is None:
        return None
    value = value.strip()
    return None if value in _NA_VALUES else value


def _parse_date(value: str | None) -> str:
    if value is None:
        raise ValueError("missing date")
    return datetime.strptime(value, "%m/%d/%Y").date().isoformat()


def _norm_zip(value: str | None) -> str | None:
    """Strip a trailing '.0' (float-parsed zips), left-pad to 5 with zeros,
    else None if it doesn't look like a 5-digit zip."""
    if value is None:
        return None
    value = value.removesuffix(".0")
    if value.isdigit() and len(value) <= 5:
        value = value.zfill(5)
    return value if _ZIP_RE.fullmatch(value) else None
